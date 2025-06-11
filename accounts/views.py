# accounts/views.py
import random
import requests
from functools import wraps
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.contrib import messages
from .models import Order, OrderItem, Medication, Supplier, AuditLog

from django.http import HttpResponse
from reportlab.pdfgen import canvas
from io import BytesIO
import json
import openpyxl
from django.utils import timezone
from datetime import timedelta
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_protect
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.utils import timezone
from datetime import datetime, timedelta
from .models import (CustomUser, UserSession, PatientRecord, 
                    Prescription, MedicalExam, AuditLog)

# =============================================
# UTILITAIRES ET VUES COMMUNES
# =============================================

MFA_CODE_VALID_MINUTES = 5

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def role_required(*roles):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if request.user.role not in roles:
                messages.error(request, "Vous n'avez pas la permission d'accéder à cette page.")
                return redirect('dashboard')
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator

def home(request):
    return render(request, 'accounts/home.html')

@csrf_protect
def user_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        role = request.POST.get('role')

        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            if user.role == role:
                mfa_code = str(random.randint(100000, 999999))
                user.mfa_code = mfa_code
                user.mfa_code_created_at = datetime.now()
                user.save()

                print(f"CODE MFA POUR {user.username}: {mfa_code}")
                
                request.session['mfa_user_id'] = user.id
                return redirect('mfa_verification')
            else:
                messages.error(request, f"Rôle incorrect. Votre rôle: {user.get_role_display()}")
        else:
            messages.error(request, "Nom d'utilisateur ou mot de passe incorrect")
    
    return render(request, 'accounts/login.html')

@csrf_protect
def mfa_verification(request):
    user_id = request.session.get('mfa_user_id')
    if not user_id:
        return redirect('login')
    
    if request.method == 'POST':
        entered_code = request.POST.get('mfa_code')
        
        try:
            user = CustomUser.objects.get(id=user_id)
            
            if user.mfa_code_created_at and (timezone.now() - user.mfa_code_created_at > timedelta(minutes=5)):
                messages.error(request, "Code expiré. Veuillez vous reconnecter.")
                return redirect('login')
            
            if user.mfa_code == entered_code:
                user.mfa_code = None
                user.mfa_code_created_at = None
                user.save()
                login(request, user)
                
                UserSession.objects.create(
                    user=user,
                    session_key=request.session.session_key,
                    is_active=True
                )
                
                AuditLog.objects.create(
                    user=user,
                    action='LOGIN',
                    model='CustomUser',
                    object_id=user.id,
                    details=f"Connexion réussie via MFA",
                    ip_address=get_client_ip(request)
                )
                
                return redirect('dashboard')
            else:
                messages.error(request, "Code incorrect")
        except CustomUser.DoesNotExist:
            messages.error(request, "Session expirée")
    
    return render(request, 'accounts/mfa_verification.html')

@login_required
def user_logout(request):
    if request.user.is_authenticated:
        UserSession.objects.filter(
            user=request.user,
            session_key=request.session.session_key
        ).update(is_active=False, logout_at=timezone.now())
        
        AuditLog.objects.create(
            user=request.user,
            action='LOGOUT',
            model='CustomUser',
            object_id=request.user.id,
            details=f"Déconnexion de l'utilisateur",
            ip_address=get_client_ip(request)
        )
        
        logout(request)
        messages.success(request, "Déconnexion réussie")
    
    return redirect('home')

# =============================================
# TABLEAU DE BORD PRINCIPAL
# =============================================

@login_required
def dashboard(request):
    """Tableau de bord principal avec redirection vers les bonnes vues"""
    role_data = {
        'DOCTOR': doctor_dashboard_data(request.user),
        'NURSE': nurse_dashboard_data(request.user),
        'PHARMACIST': pharmacist_dashboard_data(request.user),
        'ADMIN': admin_dashboard_data(request.user),
        'PATIENT': patient_dashboard_data(request.user),
    }.get(request.user.role, {})
    
    context = {
        'title': role_data.get('title', 'Tableau de Bord'),
        'cards': role_data.get('cards', []),
        'user': request.user,
        'role_display': {
            'DOCTOR': '👨‍⚕️ Médecin',
            'NURSE': '👩‍⚕️ Infirmier',
            'PHARMACIST': '💊 Pharmacien',
            'ADMIN': '👔 Administrateur',
            'PATIENT': '👤 Patient'
        }.get(request.user.role, 'Utilisateur'),
        'last_login': request.user.last_login
    }
    return render(request, 'accounts/dashboard.html', context)

# Fonctions helpers pour les tableaux de bord
def doctor_dashboard_data(user):
    return {
        'title': 'Tableau de Bord Médecin',
        'stats': {
            'patients': CustomUser.objects.filter(role='PATIENT').count(),
            'prescriptions': Prescription.objects.filter(doctor=user).count()
        },
        'cards': [
            {
                'title': 'Gestion des Patients',
                'buttons': [
                    {'url': 'create_patient_record', 'text': 'Créer un dossier', 'class': 'btn-secondary'},
                    {'url': 'view_patient_records', 'text': 'Consulter un dossier', 'class': 'btn-secondary'},
                ]
            },
            {
                'title': 'Prescriptions',
                'buttons': [
                    {'url': 'create_prescription', 'text': 'Nouvelle ordonnance', 'class': 'btn-secondary'},
                    {'url': 'view_prescriptions', 'text': 'Historique des prescriptions', 'class': 'btn-primary'},
                ]
            },
            {
                'title': 'Examens Médicaux',
                'buttons': [
                    {'url': 'request_exam', 'text': 'Demander un examen', 'class': 'btn-secondary'},
                    {'url': 'view_exam_results', 'text': 'Résultats d\'examens', 'class': 'btn-primary'},
                ]
            },
            {
                'title': 'Urgences',
                'buttons': [
                    {'url': 'emergency_access', 'text': 'Accès Urgence', 'class': 'btn-emergency'},
                    {'url': 'critical_patients', 'text': 'Patients Critiques', 'class': 'btn-emergency'},
                ]
            }
        ]
    }

def nurse_dashboard_data(user):
    return {
        'title': 'Tableau de Bord Infirmier',
        'stats': {
            'vitals_today': PatientRecord.objects.filter(last_updated__date=timezone.now().date()).count()
        },
        'cards': [
            {
                'title': 'Soins aux Patients',
                'buttons': [
                    {'url': 'record_care', 'text': 'Enregistrer des soins', 'class': 'btn-secondary'},
                    {'url': 'view_patient_vitals', 'text': 'Constantes vitales', 'class': 'btn-primary'},
                ]
            },
            {
                'title': 'Prescriptions',
                'buttons': [
                    {'url': 'view_prescriptions', 'text': 'Voir les prescriptions', 'class': 'btn-secondary'},
                    {'url': 'administer_medication', 'text': 'Administrer médicament', 'class': 'btn-primary'},
                ]
            },
            {
                'title': 'Actions Rapides',
                'buttons': [
                    # Utiliser l'URL existante view_patient_vitals
                    {'url': 'view_patient_vitals', 'text': 'Saisir Constantes', 'class': 'btn-primary'},
                    {'url': 'medication_administration', 'text': 'Administration Médicaments', 'class': 'btn-primary'},
                ]
            }
        ]
    }

def pharmacist_dashboard_data(user):
    return {
        'title': 'Tableau de Bord Pharmacien',
        'stats': {
            'pending_prescriptions': Prescription.objects.filter(is_dispensed=False).count(),
            'low_stock': 5  # À remplacer par une vraie requête
        },
        'cards': [
            {
                'title': 'Gestion des Médicaments',
                'buttons': [
                    {'url': 'view_prescriptions', 'text': 'Voir les prescriptions', 'class': 'btn-secondary'},
                    {'url': 'dispense_medication', 'text': 'Délivrer médicament', 'class': 'btn-primary'},
                ]
            },
            {
                'title': 'Stock Pharmacie',
                'buttons': [
                    {'url': 'view_inventory', 'text': 'Voir le stock', 'class': 'btn-secondary'},
                    {'url': 'request_reorder', 'text': 'Commander médicaments', 'class': 'btn-primary'},
                ]
            }
        ]
    }

def admin_dashboard_data(user):
    last_login = user.last_login.strftime("%d/%m/%Y %H:%M") if user.last_login else "Jamais"
    
    return {
        'title': 'Tableau de Bord Administrateur',
        'user_info': {
            'last_login': last_login,
            'role': 'Administrateur'
        },
        'cards': [
            {
                'title': 'Gestion Utilisateurs',
                'buttons': [
                    {
                        'url': 'admin:accounts_customuser_changelist',  # URL de la liste des users dans l'admin
                        'text': 'Gérer les utilisateurs', 
                        'class': 'btn-primary',
                        'icon': 'fas fa-users'
                    }
                ]
            },
            {
                'title': 'Journal d\'activité',
                'buttons': [
                    {
                        'url': 'admin:accounts_auditlog_changelist',  # URL des logs dans l'admin
                        'text': 'Voir les journaux', 
                        'class': 'btn-primary',
                        'icon': 'fas fa-clipboard-list'
                    }
                ]
            },
            {
                'title': 'Configuration',
                'buttons': [
                    {
                        'url': 'admin:accounts_securitysettings_changelist', 
                        'text': 'Paramètres de sécurité', 
                        'class': 'btn-secondary',
                        'icon': 'fas fa-lock',
                        'description': 'Configurer les règles de sécurité et permissions'
                    },
                    {
                        'url': 'admin:accounts_systemconfig_changelist',
                        'text': 'Configuration système', 
                        'class': 'btn-secondary',
                        'icon': 'fas fa-cog',
                        'description': 'Paramètres globaux de l\'application'
                    }
                ]
            }
        ]
    }
def patient_dashboard_data(user):
    return {
        'title': 'Mon Espace Patient',
        'stats': {
            'prescriptions': Prescription.objects.filter(patient=user).count(),
            'appointments': 0  # À remplacer par une vraie requête
        },
        'cards': [
            {
                'title': 'Mon Dossier Médical',
                'buttons': [
                    {'url': 'view_medical_record', 'text': 'Consulter mon dossier', 'class': 'btn-secondary'},
                    {'url': 'download_medical_record', 'text': 'Télécharger mon dossier', 'class': 'btn-primary'},
                ]
            },
            {
                'title': 'Mes Rendez-vous',
                'buttons': [
                    {'url': 'view_appointments', 'text': 'Mes rendez-vous', 'class': 'btn-secondary'},
                    {'url': 'book_appointment', 'text': 'Prendre rendez-vous', 'class': 'btn-primary'},
                ]
            }
        ]
    }

# =============================================
# VUES MÉDECINS
# =============================================

@login_required
@role_required('DOCTOR')
def create_patient_record(request):
    if request.method == 'POST':
        try:
            patient = CustomUser.objects.create_user(
                username=request.POST.get('username'),
                password=request.POST.get('password'),
                first_name=request.POST.get('first_name'),
                last_name=request.POST.get('last_name'),
                phone_number=request.POST.get('phone_number'),
                role='PATIENT',
                is_verified=True
            )
            
            PatientRecord.objects.create(
                patient=patient,
                created_by=request.user,
                medical_history=request.POST.get('medical_history'),
                allergies=request.POST.get('allergies'),
                current_medications=request.POST.get('current_medications')
            )
            
            AuditLog.objects.create(
                user=request.user,
                action='CREATE',
                model='PatientRecord',
                object_id=patient.id,
                details=f"Création dossier patient pour {patient.get_full_name()}",
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, "Patient créé avec succès!")
            return redirect('view_patient_records')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la création: {str(e)}")
    
    return render(request, 'accounts/doctor/create_patient_record.html')

@login_required
@role_required('DOCTOR')
def view_patient_records(request):
    patients = CustomUser.objects.filter(role='PATIENT')
    records = PatientRecord.objects.select_related('patient').all()
    return render(request, 'accounts/doctor/view_patient_records.html', {
        'patients': patients,
        'records': records
    })

@login_required
@role_required('DOCTOR')
def view_patient_detail(request, patient_id):
    patient = get_object_or_404(CustomUser, id=patient_id, role='PATIENT')
    record = get_object_or_404(PatientRecord, patient=patient)
    prescriptions = Prescription.objects.filter(patient=patient)
    exams = MedicalExam.objects.filter(patient=patient)
    
    return render(request, 'accounts/doctor/view_patient_detail.html', {
        'patient': patient,
        'record': record,
        'prescriptions': prescriptions,
        'exams': exams
    })

@login_required
@role_required('DOCTOR')
def create_prescription(request):
    if request.method == 'POST':
        try:
            patient_id = request.POST.get('patient_id')
            patient = get_object_or_404(CustomUser, id=patient_id, role='PATIENT')
            
            prescription = Prescription.objects.create(
                patient=patient,
                doctor=request.user,
                medication=request.POST.get('medication'),
                dosage=request.POST.get('dosage'),
                instructions=request.POST.get('instructions')
            )
            
            record, created = PatientRecord.objects.get_or_create(patient=patient)
            if not created:
                record.current_medications += f"\n{prescription.medication} ({prescription.dosage})"
                record.save()
            
            messages.success(request, "Prescription créée avec succès!")
            return redirect('view_prescriptions')
            
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
    
    patients = CustomUser.objects.filter(role='PATIENT')
    return render(request, 'accounts/doctor/create_prescription.html', {
        'patients': patients,
        'medications': ['Paracétamol', 'Ibuprofène', 'Amoxicilline']
    })

@login_required
@role_required('PHARMACIST', 'DOCTOR', 'NURSE') 
def view_prescriptions(request):
    if request.user.role == 'PHARMACIST':
        prescriptions = Prescription.objects.all().select_related('patient', 'doctor')
    else:
        prescriptions = Prescription.objects.filter(doctor=request.user).select_related('patient')
    return render(request, 'accounts/doctor/view_prescriptions.html', {
        'prescriptions': prescriptions
    })

@login_required
@role_required('DOCTOR')
def request_exam(request):
    if request.method == 'POST':
        try:
            patient_id = request.POST.get('patient_id')
            patient = get_object_or_404(CustomUser, id=patient_id, role='PATIENT')
            
            exam = MedicalExam.objects.create(
                patient=patient,
                requested_by=request.user,
                exam_type=request.POST.get('exam_type'),
                notes=request.POST.get('notes')
            )
            
            AuditLog.objects.create(
                user=request.user,
                action='CREATE',
                model='MedicalExam',
                object_id=exam.id,
                details=f"Examen {exam.get_exam_type_display()} demandé pour {patient.get_full_name()}",
                ip_address=get_client_ip(request)
            )
            
            record, created = PatientRecord.objects.get_or_create(patient=patient)
            if not created:
                record.medical_history += f"\nExamen {exam.get_exam_type_display()} demandé le {timezone.now().strftime('%d/%m/%Y')}"
                record.save()
            
            messages.success(request, "Demande d'examen enregistrée!")
            return redirect('view_exam_results')
            
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
    
    patients = CustomUser.objects.filter(role='PATIENT')
    return render(request, 'accounts/doctor/request_exam.html', {
        'patients': patients,
        'exam_types': MedicalExam.EXAM_TYPES
    })

@login_required
@role_required('DOCTOR')
def view_exam_results(request):
    exams = MedicalExam.objects.filter(requested_by=request.user).select_related('patient')
    return render(request, 'accounts/doctor/view_exam_results.html', {
        'exams': exams
    })

@login_required
@role_required('DOCTOR')
def emergency_access(request):
    critical_records = PatientRecord.objects.filter(
        is_critical=True
    ).select_related('patient')
    
    AuditLog.objects.create(
        user=request.user,
        action='ACCESS',
        model='Emergency',
        object_id=0,
        details="Accès au mode urgence activé",
        ip_address=get_client_ip(request)
    )
    
    return render(request, 'accounts/doctor/emergency_access.html', {
        'critical_patients': [record.patient for record in critical_records],
        'critical_count': critical_records.count()
    })
@login_required
@role_required('DOCTOR')
def critical_patients(request):
    if request.method == 'POST':
        patient_id = request.POST.get('patient_id')
        action = request.POST.get('action')
        
        patient = get_object_or_404(CustomUser, id=patient_id, role='PATIENT')
        record = PatientRecord.objects.get(patient=patient)
        
        if action == 'mark_critical':
            record.is_critical = True
            record.critical_since = timezone.now()
            record.priority = request.POST.get('priority', '2')
            record.critical_reason = request.POST.get('reason', '')
            messages.success(request, f"{patient.get_full_name()} marqué comme patient critique")
        elif action == 'unmark_critical':
            record.is_critical = False
            record.critical_since = None
            record.priority = ''
            messages.success(request, f"{patient.get_full_name()} retiré des patients critiques")
        
        record.save()
        return redirect('critical_patients')
    
    # Récupération optimisée des patients critiques
    critical_records = PatientRecord.objects.filter(
        is_critical=True
    ).select_related('patient')
    
    # Récupération des patients non critiques
    regular_patients = CustomUser.objects.filter(
        role='PATIENT'
    ).exclude(
        id__in=[record.patient_id for record in critical_records]
    )
    
    return render(request, 'accounts/doctor/critical_patients.html', {
        'critical_patients': [record.patient for record in critical_records],
        'regular_patients': regular_patients,
        'critical_count': critical_records.count()
    })

@login_required
@role_required('DOCTOR')
def protected_records(request):
    critical_patients = PatientRecord.objects.filter(
        medical_history__icontains='urgence'
    ).select_related('patient')
    return render(request, 'accounts/doctor/protected_records.html', {
        'critical_patients': critical_patients
    })

# Ajoutez cette vue dans votre fichier views.py

# Ajoutez ces vues dans votre fichier views.py

@login_required
@role_required('DOCTOR', 'PHARMACIST')
def print_prescription(request, prescription_id):
    """Vue pour imprimer/afficher une prescription"""
    prescription = get_object_or_404(Prescription, id=prescription_id)
    
    # Vérifier que le médecin peut accéder à cette prescription
    if request.user.role == 'DOCTOR' and prescription.doctor != request.user:
        messages.error(request, "Vous n'avez pas accès à cette prescription.")
        return redirect('view_prescriptions')
    
    # Logger l'accès à la prescription
    AuditLog.objects.create(
        user=request.user,
        action='VIEW',
        model='Prescription',
        object_id=prescription.id,
        details=f"Impression prescription pour {prescription.patient.get_full_name()}",
        ip_address=get_client_ip(request)
    )
    
    context = {
        'prescription': prescription,
        'patient': prescription.patient,
        'doctor': prescription.doctor,
    }
    
    return render(request, 'accounts/doctor/print_prescription.html', context)

@login_required
@role_required('DOCTOR')
def edit_prescription(request, prescription_id):
    """Vue pour modifier une prescription"""
    prescription = get_object_or_404(Prescription, id=prescription_id)
    
    # Vérifier que le médecin peut modifier cette prescription
    if prescription.doctor != request.user:
        messages.error(request, "Vous ne pouvez modifier que vos propres prescriptions.")
        return redirect('view_prescriptions')
    
    if request.method == 'POST':
        try:
            # Sauvegarder les anciennes valeurs pour l'audit
            old_medication = prescription.medication
            old_dosage = prescription.dosage
            old_instructions = prescription.instructions
            
            # Modifier la prescription
            prescription.medication = request.POST.get('medication')
            prescription.dosage = request.POST.get('dosage')
            prescription.instructions = request.POST.get('instructions')
            prescription.save()
            
            # Logger la modification
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                model='Prescription',
                object_id=prescription.id,
                details=f"Prescription modifiée pour {prescription.patient.get_full_name()}: {old_medication} -> {prescription.medication}",
                ip_address=get_client_ip(request)
            )
            
            # Mettre à jour le dossier patient
            record = PatientRecord.objects.get(patient=prescription.patient)
            # Remplacer l'ancienne médication par la nouvelle dans l'historique
            if old_medication in record.current_medications:
                record.current_medications = record.current_medications.replace(
                    f"{old_medication} ({old_dosage})",
                    f"{prescription.medication} ({prescription.dosage})"
                )
                record.save()
            
            messages.success(request, "Prescription modifiée avec succès!")
            return redirect('view_prescriptions')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la modification: {str(e)}")
    
    # Récupérer la liste des patients et médicaments pour le formulaire
    patients = CustomUser.objects.filter(role='PATIENT')
    medications = ['Paracétamol', 'Ibuprofène', 'Amoxicilline', 'Aspirine', 'Doliprane', 'Advil']
    
    context = {
        'prescription': prescription,
        'patients': patients,
        'medications': medications
    }
    
    return render(request, 'accounts/doctor/edit_prescription.html', context)

@login_required
@role_required('DOCTOR')
def delete_prescription(request, prescription_id):
    """Vue pour supprimer une prescription"""
    prescription = get_object_or_404(Prescription, id=prescription_id)
    
    # Vérifier que le médecin peut supprimer cette prescription
    if prescription.doctor != request.user:
        messages.error(request, "Vous ne pouvez supprimer que vos propres prescriptions.")
        return redirect('view_prescriptions')
    
    if request.method == 'POST':
        # Logger la suppression
        AuditLog.objects.create(
            user=request.user,
            action='DELETE',
            model='Prescription',
            object_id=prescription.id,
            details=f"Prescription supprimée pour {prescription.patient.get_full_name()}: {prescription.medication}",
            ip_address=get_client_ip(request)
        )
        
        # Retirer la médication du dossier patient
        try:
            record = PatientRecord.objects.get(patient=prescription.patient)
            medication_entry = f"{prescription.medication} ({prescription.dosage})"
            if medication_entry in record.current_medications:
                record.current_medications = record.current_medications.replace(
                    f"\n{medication_entry}", ""
                ).replace(medication_entry, "")
                record.save()
        except PatientRecord.DoesNotExist:
            pass
        
        prescription.delete()
        messages.success(request, "Prescription supprimée avec succès!")
        return redirect('view_prescriptions')
    
    return render(request, 'accounts/doctor/confirm_delete_prescription.html', {
        'prescription': prescription
    })

# Ajoutez ces vues dans votre fichier views.py

@login_required
@role_required('DOCTOR')
def view_exam_detail(request, exam_id):
    """Vue pour voir les détails d'un examen médical"""
    exam = get_object_or_404(MedicalExam, id=exam_id)
    
    # Vérifier que le médecin peut accéder à cet examen
    if exam.requested_by != request.user:
        messages.error(request, "Vous n'avez pas accès à cet examen.")
        return redirect('view_exam_results')
    
    # Logger l'accès à l'examen
    AuditLog.objects.create(
        user=request.user,
        action='VIEW',
        model='MedicalExam',
        object_id=exam.id,
        details=f"Consultation détail examen {exam.get_exam_type_display()} pour {exam.patient.get_full_name()}",
        ip_address=get_client_ip(request)
    )
    
    context = {
        'exam': exam,
        'patient': exam.patient,
    }
    
    return render(request, 'accounts/doctor/view_exam_detail.html', context)

@login_required
@role_required('DOCTOR')
def edit_exam(request, exam_id):
    """Vue pour modifier un examen médical"""
    exam = get_object_or_404(MedicalExam, id=exam_id)
    
    # Vérifier que le médecin peut modifier cet examen
    if exam.requested_by != request.user:
        messages.error(request, "Vous ne pouvez modifier que vos propres demandes d'examen.")
        return redirect('view_exam_results')
    
    if request.method == 'POST':
        try:
            # Sauvegarder les anciennes valeurs pour l'audit
            old_exam_type = exam.exam_type
            old_notes = exam.notes
            
            # Modifier l'examen
            exam.exam_type = request.POST.get('exam_type')
            exam.notes = request.POST.get('notes')
            
            # Si des résultats sont fournis, les ajouter
            if request.POST.get('results'):
                exam.results = request.POST.get('results')
                exam.status = 'COMPLETED'
                exam.completed_at = timezone.now()
            
            exam.save()
            
            # Logger la modification
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                model='MedicalExam',
                object_id=exam.id,
                details=f"Examen modifié pour {exam.patient.get_full_name()}: {old_exam_type} -> {exam.exam_type}",
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, "Examen modifié avec succès!")
            return redirect('view_exam_results')
            
        except Exception as e:
            messages.error(request, f"Erreur lors de la modification: {str(e)}")
    
    patients = CustomUser.objects.filter(role='PATIENT')
    
    context = {
        'exam': exam,
        'patients': patients,
        'exam_types': MedicalExam.EXAM_TYPES
    }
    
    return render(request, 'accounts/doctor/edit_exam.html', context)

@login_required
@role_required('DOCTOR')
def delete_exam(request, exam_id):
    """Vue pour supprimer une demande d'examen"""
    exam = get_object_or_404(MedicalExam, id=exam_id)
    
    # Vérifier que le médecin peut supprimer cet examen
    if exam.requested_by != request.user:
        messages.error(request, "Vous ne pouvez supprimer que vos propres demandes d'examen.")
        return redirect('view_exam_results')
    
    if request.method == 'POST':
        # Logger la suppression
        AuditLog.objects.create(
            user=request.user,
            action='DELETE',
            model='MedicalExam',
            object_id=exam.id,
            details=f"Demande d'examen supprimée pour {exam.patient.get_full_name()}: {exam.get_exam_type_display()}",
            ip_address=get_client_ip(request)
        )
        
        exam.delete()
        messages.success(request, "Demande d'examen supprimée avec succès!")
        return redirect('view_exam_results')
    
    return render(request, 'accounts/doctor/confirm_delete_exam.html', {
        'exam': exam
    })

@login_required
@role_required('DOCTOR')
def add_exam_results(request, exam_id):
    """Vue pour ajouter les résultats d'un examen"""
    exam = get_object_or_404(MedicalExam, id=exam_id)
    
    # Vérifier que le médecin peut modifier cet examen
    if exam.requested_by != request.user:
        messages.error(request, "Vous n'avez pas accès à cet examen.")
        return redirect('view_exam_results')
    
    if request.method == 'POST':
        try:
            exam.results = request.POST.get('results')
            exam.status = 'COMPLETED'
            exam.completed_at = timezone.now()
            exam.save()
            
            # Mettre à jour l'historique médical du patient
            record, created = PatientRecord.objects.get_or_create(patient=exam.patient)
            record.medical_history += f"\n{exam.get_exam_type_display()} - {timezone.now().strftime('%d/%m/%Y')}: {exam.results}"
            record.save()
            
            # Logger l'ajout des résultats
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                model='MedicalExam',
                object_id=exam.id,
                details=f"Résultats ajoutés pour l'examen {exam.get_exam_type_display()} de {exam.patient.get_full_name()}",
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, "Résultats de l'examen ajoutés avec succès!")
            return redirect('view_exam_results')
            
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
    
    return render(request, 'accounts/doctor/add_exam_results.html', {
        'exam': exam
    })

# =============================================
# VUES INFIRMIERS
# =============================================

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.utils import timezone
from datetime import datetime, timedelta
from .models import CustomUser, PatientRecord, Prescription, MedicalExam, AuditLog, Appointment
from .decorators import role_required

def get_client_ip(request):
    """Helper function to get client IP address"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

@login_required
@role_required('NURSE')
def record_care(request):
    # Récupérer tous les patients
    patients = CustomUser.objects.filter(role='PATIENT').order_by('last_name', 'first_name')
    
    if request.method == 'POST':
        try:
            patient_id = request.POST.get('patient_id')
            if not patient_id:
                raise ValueError("Patient non spécifié")
                
            patient = get_object_or_404(CustomUser, id=patient_id, role='PATIENT')
            
            care_type = request.POST.get('care_type')
            care_time = request.POST.get('care_time')
            care_description = request.POST.get('care_description')
            observations = request.POST.get('observations')
            pain_level = request.POST.get('pain_level')
            medication_given = request.POST.get('medication_given')
            
            record, created = PatientRecord.objects.get_or_create(patient=patient)
            record.medical_history += f"\n\nSoins du {timezone.now().strftime('%d/%m/%Y %H:%M')}:\n"
            record.medical_history += f"Type: {care_type}\n"
            record.medical_history += f"Description: {care_description}\n"
            record.medical_history += f"Observations: {observations}\n"
            record.medical_history += f"Niveau de douleur: {pain_level}/10\n"
            if medication_given:
                record.medical_history += f"Médicament administré: {medication_given}\n"
            record.save()
            
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                model='PatientRecord',
                object_id=record.id,
                details=f"Soins enregistrés pour {patient.get_full_name()}",
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, "Soins enregistrés avec succès!")
            return redirect('view_patient_vitals', patient_id=patient.id)
            
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
            return redirect('record_care')
    
    return render(request, 'accounts/nurse/record_care.html', {
        'patients': patients
    })

@login_required
@role_required('NURSE')
def view_patient_vitals(request, patient_id=None):
    if patient_id is None:
        patients = CustomUser.objects.filter(role='PATIENT')
        return render(request, 'accounts/nurse/select_patient.html', {'patients': patients})
    
    patient = get_object_or_404(CustomUser, id=patient_id, role='PATIENT')
    record = get_object_or_404(PatientRecord, patient=patient)
    
    # Extraire les dernières constantes vitales de l'historique médical
    last_vitals = {
        'temperature': None,
        'heart_rate': None,
        'systolic': None,
        'diastolic': None,
        'oxygen': None,
        'time': None
    }
    
    # Analyse simplifiée de l'historique pour trouver les dernières constantes
    history_lines = record.medical_history.split('\n')
    for line in reversed(history_lines):
        if 'TA:' in line:
            parts = line.split(',')
            for part in parts:
                if 'TA:' in part:
                    try:
                        bp_values = part.split('TA:')[1].strip().split('/')
                        if len(bp_values) == 2:
                            last_vitals['systolic'], last_vitals['diastolic'] = bp_values
                    except (IndexError, ValueError):
                        pass
                elif 'Temp:' in part:
                    try:
                        last_vitals['temperature'] = part.split('Temp:')[1].strip().split('°')[0]
                    except (IndexError, ValueError):
                        pass
                elif 'Pouls:' in part:
                    try:
                        last_vitals['heart_rate'] = part.split('Pouls:')[1].strip().split(' ')[0]
                    except (IndexError, ValueError):
                        pass
                elif 'O₂:' in part:
                    try:
                        last_vitals['oxygen'] = part.split('O₂:')[1].strip().split('%')[0]
                    except (IndexError, ValueError):
                        pass
                elif 'Constantes du' in part:
                    try:
                        last_vitals['time'] = part.split('Constantes du')[1].strip()
                    except (IndexError, ValueError):
                        pass
            break  # Stop after finding the first (most recent) vitals entry
    
    return render(request, 'accounts/nurse/view_patient_vitals.html', {
        'patient': patient,
        'record': record,
        'last_vitals': last_vitals
    })

@login_required
@role_required('NURSE')
def quick_vitals(request, patient_id):
    patient = get_object_or_404(CustomUser, id=patient_id, role='PATIENT')
    
    if request.method == 'POST':
        try:
            temperature = request.POST.get('temperature')
            heart_rate = request.POST.get('heart_rate')
            systolic = request.POST.get('systolic')
            diastolic = request.POST.get('diastolic')
            oxygen = request.POST.get('oxygen')
            respiratory_rate = request.POST.get('respiratory_rate')
            notes = request.POST.get('notes')
            
            record = PatientRecord.objects.get(patient=patient)
            record.medical_history += f"\n\nConstantes du {timezone.now().strftime('%d/%m/%Y %H:%M')}:\n"
            record.medical_history += f"TA: {systolic}/{diastolic}, "
            record.medical_history += f"Temp: {temperature}°C, "
            record.medical_history += f"Pouls: {heart_rate} bpm, "
            record.medical_history += f"O₂: {oxygen}%, "
            record.medical_history += f"FR: {respiratory_rate}/min"
            if notes:
                record.medical_history += f"\nNotes: {notes}"
            record.save()
            
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                model='PatientRecord',
                object_id=record.id,
                details=f"Constantes vitales enregistrées pour {patient.get_full_name()}",
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, "Constantes enregistrées!")
            return redirect('view_patient_vitals', patient_id=patient.id)
            
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
    
    return render(request, 'accounts/nurse/quick_vitals.html', {
        'patient': patient
    })

@login_required
@role_required('NURSE')
def administer_medication(request, patient_id=None):  # AJOUT DU PARAMETRE patient_id=None
    # Récupérer patient_id depuis l'URL ou les paramètres GET
    if not patient_id:
        patient_id = request.GET.get('patient_id')
    
    patients = CustomUser.objects.filter(role='PATIENT')
    
    # Si aucun patient_id et qu'il y a des patients, rediriger vers le premier
    if not patient_id and patients.exists():
        return redirect('administer_medication', patient_id=patients.first().id)
    
    patient = get_object_or_404(CustomUser, id=patient_id, role='PATIENT') if patient_id else None
    prescriptions = Prescription.objects.filter(patient=patient, is_dispensed=True) if patient else []
    
    if request.method == 'POST':
        try:
            prescription_id = request.POST.get('prescription_id')
            prescription = get_object_or_404(Prescription, id=prescription_id)
            
            prescription.instructions += f"\n\nAdministré le {timezone.now().strftime('%d/%m/%Y %H:%M')} par {request.user.get_full_name()}"
            prescription.is_administered = True
            prescription.administered_at = timezone.now()
            prescription.save()
            
            record = PatientRecord.objects.get(patient=prescription.patient)
            record.medical_history += f"\nMédicament administré: {prescription.medication} ({prescription.dosage}) à {timezone.now().strftime('%H:%M')}"
            record.save()
            
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                model='Prescription',
                object_id=prescription.id,
                details=f"Médicament administré à {prescription.patient.get_full_name()}",
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, "Médicament administré avec succès!")
            return redirect('administer_medication', patient_id=prescription.patient.id)
            
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
    
    return render(request, 'accounts/nurse/administer_medication.html', {
        'patients': patients,
        'patient': patient,
        'prescriptions': prescriptions
    })

@login_required
@role_required('NURSE')
def medication_administration(request):
    prescriptions = Prescription.objects.filter(is_dispensed=True).select_related('patient')
    return render(request, 'accounts/nurse/medication_administration.html', {
        'prescriptions': prescriptions
    })
def select_patient_for_vitals(request):
    """Vue pour sélectionner un patient avant de saisir les constantes vitales"""
    if not request.user.is_authenticated or request.user.role != 'NURSE':
        return redirect('login')
    
    # Récupérer tous les utilisateurs ayant le rôle PATIENT
    patients = CustomUser.objects.filter(role='PATIENT').order_by('last_name', 'first_name')
    
    context = {
        'patients': patients,
        'title': 'Sélectionner un Patient - Constantes Vitales'
    }
    return render(request, 'accounts/nurse/select_patient_vitals.html', context)

# =============================================
# VUES PHARMACIENS
# =============================================

@login_required
@role_required('PHARMACIST')
def dispense_medication(request):
    if request.method == 'POST':
        try:
            prescription_id = request.POST.get('prescription_id')
            prescription = get_object_or_404(Prescription, id=prescription_id)
            
            prescription.is_dispensed = True
            prescription.dispensed_by = request.user
            prescription.dispensed_at = timezone.now()
            prescription.save()
            
            # Enregistrer dans l'historique du patient
            record = PatientRecord.objects.get(patient=prescription.patient)
            record.medical_history += f"\nMédicament délivré: {prescription.medication} ({prescription.dosage}) le {timezone.now().strftime('%d/%m/%Y')} par {request.user.get_full_name()}"
            record.save()
            
            messages.success(request, "Médicament délivré avec succès!")
            return redirect('view_prescriptions')
            
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
    
    prescription_id = request.GET.get('prescription_id')
    if not prescription_id:
        return redirect('view_prescriptions')
    
    prescription = get_object_or_404(
        Prescription.objects.select_related('patient', 'doctor'), 
        id=prescription_id
    )
              
    AuditLog.objects.create(
        user=request.user,
        action='UPDATE',
        model='Prescription',
        object_id=prescription.id,
        details=f"Médicament {prescription.medication} délivré à {prescription.patient.get_full_name()}",
        ip_address=get_client_ip(request)
         )
    return render(request, 'accounts/pharmacist/dispense_medication.html', {
        'prescription': prescription
    })
@login_required
@role_required('PHARMACIST')
def process_order(request):
    if request.method == 'POST':
        try:
            order_data = json.loads(request.POST.get('order_data'))
            # Traiter la commande ici
            
            # Log d'audit
            AuditLog.objects.create(
                user=request.user,
                action='CREATE',
                model='Order',
                object_id=0,  # ou l'ID de la commande créée
                details=f"Commande de médicaments passée",
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, "Commande passée avec succès!")
            return redirect('view_inventory')
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
            return redirect('request_reorder')

@login_required
@role_required('PHARMACIST')
def process_order(request):
    if request.method == 'POST':
        try:
            # Récupérer les données JSON
            order_data_str = request.POST.get('order_data', '[]')
            order_items = json.loads(order_data_str)
            
            # Récupérer les autres données du formulaire
            supplier_id = request.POST.get('supplier')
            notes = request.POST.get('notes', '')
            
            # Validation
            if not order_items:
                messages.error(request, 'Aucun médicament sélectionné pour la commande.')
                return redirect('request_reorder')
            
            if not supplier_id:
                messages.error(request, 'Veuillez sélectionner un fournisseur.')
                return redirect('request_reorder')
            
            # Récupérer le fournisseur
            supplier = get_object_or_404(Supplier, id=supplier_id)
            
            # Créer la commande
            order = Order.objects.create(
                pharmacist=request.user,
                supplier=supplier,
                notes=notes,
                status='PENDING'
            )
            
            # Ajouter les items de la commande
            for item in order_items:
                medication = get_object_or_404(Medication, id=item['medId'])
                OrderItem.objects.create(
                    order=order,
                    medication=medication,
                    quantity=item['quantity'],
                    unit_price=item.get('unitPrice', medication.unit_price)
                )
            
            # Log d'audit
            AuditLog.objects.create(
                user=request.user,
                action='CREATE',
                model='Order',
                object_id=order.id,
                details=f"Commande #{order.id} passée auprès de {supplier.name}",
                ip_address=get_client_ip(request)
            )
            
            messages.success(request, f"Commande #{order.id} passée avec succès!")
            return redirect('view_orders')
            
        except json.JSONDecodeError:
            messages.error(request, 'Erreur lors du traitement des données de commande.')
            return redirect('request_reorder')
        except Exception as e:
            messages.error(request, f"Erreur lors du traitement de la commande: {str(e)}")
            return redirect('request_reorder')
    
    return redirect('request_reorder')


@login_required
@role_required('PHARMACIST')
def view_orders(request):
    orders = Order.objects.filter(pharmacist=request.user).select_related('supplier').order_by('-order_date')
    
    # Statistiques
    total_orders = orders.count()
    pending_orders = orders.filter(status='PENDING').count()
    delivered_orders = orders.filter(status='DELIVERED').count()
    
    return render(request, 'accounts/pharmacist/view_orders.html', {
        'orders': orders,
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'delivered_orders': delivered_orders,
    })

@login_required
@role_required('PHARMACIST')
def view_orders(request):
    orders = Order.objects.filter(pharmacist=request.user).select_related('supplier').prefetch_related('items').order_by('-order_date')
    
    # Ajoutez le total à chaque commande
    for order in orders:
        order.total = sum(item.quantity * item.unit_price for item in order.items.all())
    
    # Statistiques
    total_orders = orders.count()
    pending_orders = orders.filter(status='PENDING').count()
    delivered_orders = orders.filter(status='DELIVERED').count()
    
    return render(request, 'accounts/pharmacist/view_orders.html', {
        'orders': orders,
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'delivered_orders': delivered_orders,
    })


@login_required
@role_required('PHARMACIST')
def view_order_detail(request, order_id):
    order = get_object_or_404(Order, id=order_id, pharmacist=request.user)
    order_items = order.items.all().select_related('medication')
    
    # Calcul du total
    order.total = sum(item.quantity * item.unit_price for item in order_items)
    
    return render(request, 'accounts/pharmacist/view_order_detail.html', {
        'order': order,
        'order_items': order_items,
    })
@login_required
@role_required('PHARMACIST')
def view_inventory(request):
    # Données simulées - à remplacer par vos données réelles
    inventory = [
        {'id': 1, 'name': 'Paracétamol', 'code': 'PARA500', 'dosage': '500mg', 
         'stock': 150, 'max_stock': 200, 'threshold': 50, 'supplier': 'PharmaSen'},
        {'id': 2, 'name': 'Ibuprofène', 'code': 'IBUP200', 'dosage': '200mg', 
         'stock': 80, 'max_stock': 150, 'threshold': 30, 'supplier': 'MediPlus'},
        {'id': 3, 'name': 'Amoxicilline', 'code': 'AMOX500', 'dosage': '500mg', 
         'stock': 45, 'max_stock': 100, 'threshold': 20, 'supplier': 'PharmaSen'},
    ]
    
    context = {
        'inventory': inventory,
        'total_items': len(inventory),
        'low_stock_items': len([i for i in inventory if i['stock'] <= i['threshold'] and i['stock'] > 0]),
        'out_of_stock_items': len([i for i in inventory if i['stock'] == 0]),
    }
    return render(request, 'accounts/pharmacist/view_inventory.html', context)

@login_required
@role_required('PHARMACIST')
def request_reorder(request):
    # Données simulées - à remplacer par vos données réelles
    medications = [
        {'id': 1, 'name': 'Paracétamol', 'dosage': '500mg', 'stock': 150, 'threshold': 50, 'supplier': 'PharmaSen', 'reorder_quantity': 50},
        {'id': 2, 'name': 'Ibuprofène', 'dosage': '200mg', 'stock': 80, 'threshold': 30, 'supplier': 'MediPlus', 'reorder_quantity': 70},
        {'id': 3, 'name': 'Amoxicilline', 'dosage': '500mg', 'stock': 45, 'threshold': 20, 'supplier': 'PharmaSen', 'reorder_quantity': 55},
    ]
    
    suppliers = [
        {'id': 1, 'name': 'PharmaSen', 'delivery_time': 3, 'shipping_cost': 5000},
        {'id': 2, 'name': 'MediPlus', 'delivery_time': 5, 'shipping_cost': 3000},
        {'id': 3, 'name': 'SantéAfrique', 'delivery_time': 7, 'shipping_cost': 2000},
    ]
    
    return render(request, 'accounts/pharmacist/request_reorder.html', {
        'medications': medications,
        'suppliers': suppliers
    })

# =============================================
# VUES ADMINISTRATEURS
# =============================================

@login_required
@role_required('ADMIN')
def manage_users(request):
    users = CustomUser.objects.all().order_by('-date_joined')
    return render(request, 'accounts/admin/manage_users.html', {
        'users': users
    })

@login_required
@role_required('ADMIN')
def view_audit_logs(request):
    logs = AuditLog.objects.select_related('user').order_by('-timestamp')
    
    # Filtrage
    action_filter = request.GET.get('action')
    if action_filter:
        logs = logs.filter(action=action_filter)
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(logs, 25)
    
    try:
        logs_page = paginator.page(page)
    except PageNotAnInteger:
        logs_page = paginator.page(1)
    except EmptyPage:
        logs_page = paginator.page(paginator.num_pages)
    
    return render(request, 'accounts/admin/audit_logs.html', {
        'logs': logs_page,
        'actions': AuditLog.ACTION_TYPES,
        'selected_action': action_filter
    })

# views.py
from django.contrib import messages
from .models import SystemConfig, SecuritySettings

# views.py
from django.contrib import messages
from django.utils import timezone
from .models import SystemConfig, SecuritySettings

@login_required
@role_required('ADMIN')
def system_config(request):
    config = SystemConfig.get_config()
    
    if request.method == 'POST':
        # Paramètres Généraux
        config.site_name = request.POST.get('site_name', config.site_name)
        config.timezone = request.POST.get('timezone', config.timezone)
        config.language = request.POST.get('language', config.language)
        
        # Notifications
        config.email_notifications = 'email_notifications' in request.POST
        config.sms_notifications = 'sms_notifications' in request.POST
        config.notification_email = request.POST.get('notification_email', config.notification_email)
        
        # Sauvegarde
        config.backup_frequency = request.POST.get('backup_frequency', config.backup_frequency)
        config.backup_location = request.POST.get('backup_location', config.backup_location)
        
        config.save()
        messages.success(request, "Configuration mise à jour avec succès!")
        return redirect('system_config')
    
    # Utilisez timezone.now() et les utilitaires Django plutôt que pytz
    timezones = [
        'UTC',
        'Europe/Paris',
        'America/New_York',
        'Asia/Tokyo',
        # Ajoutez d'autres fuseaux horaires selon vos besoins
    ]
    
    return render(request, 'accounts/admin/system_config.html', {
        'config': config,
        'timezones': timezones
    })

@login_required
@role_required('ADMIN')
def security_settings(request):
    settings = SecuritySettings.get_settings()
    
    if request.method == 'POST':
        # Authentification
        settings.mfa_enabled = 'mfa_enabled' in request.POST
        settings.mfa_method = request.POST.get('mfa_method', settings.mfa_method)
        
        # Verrouillage de compte
        settings.max_attempts = int(request.POST.get('max_attempts', settings.max_attempts))
        settings.lockout_time = int(request.POST.get('lockout_time', settings.lockout_time))
        
        # Politique de mot de passe
        settings.min_length = int(request.POST.get('min_length', settings.min_length))
        settings.require_complexity = request.POST.get('require_complexity', settings.require_complexity)
        settings.expiry_days = int(request.POST.get('expiry_days', settings.expiry_days))
        
        settings.save()
        messages.success(request, "Paramètres de sécurité mis à jour!")
        return redirect('security_settings')
    
    return render(request, 'admin/security_settings.html', {
        'settings': settings
    })
# =============================================
# VUES PATIENTS
# =============================================

@login_required
@role_required('PATIENT')
def view_medical_record(request):
    try:
        record = PatientRecord.objects.get(patient=request.user)
        prescriptions = Prescription.objects.filter(patient=request.user)
        exams = MedicalExam.objects.filter(patient=request.user)
        
        return render(request, 'accounts/patient/view_medical_record.html', {
            'record': record,
            'prescriptions': prescriptions,
            'exams': exams
        })
        
    except PatientRecord.DoesNotExist:
        messages.error(request, "Aucun dossier médical trouvé")
        return redirect('dashboard')

@login_required
@role_required('PATIENT')
def download_medical_record(request):
    messages.info(request, "Fonctionnalité de téléchargement en développement")
    return redirect('view_medicalrecord')

@login_required
@role_required('PATIENT')
def book_appointment(request):
    if request.method == 'POST':
        try:
            doctor_id = request.POST.get('doctor_id')
            date = request.POST.get('date')
            time = request.POST.get('time')
            reason = request.POST.get('reason')
            custom_reason = request.POST.get('custom_reason')
            notes = request.POST.get('notes')
            
            doctor = get_object_or_404(CustomUser, id=doctor_id, role='DOCTOR')
            
            appointment = Appointment.objects.create(
                patient=request.user,
                doctor=doctor,
                date=date,
                time=time,
                reason=reason,
                custom_reason=custom_reason if reason == 'OTHER' else '',
                notes=notes,
                status='SCHEDULED'
            )
            
            # Mettre à jour le dossier patient
            record = PatientRecord.objects.get(patient=request.user)
            record.next_appointment = date
            record.save()
            
            messages.success(request, "Rendez-vous confirmé avec succès!")
            return redirect('view_appointments')
            
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
    
    doctors = CustomUser.objects.filter(role='DOCTOR')
    return render(request, 'accounts/patient/book_appointment.html', {
        'doctors': doctors
    })

@login_required
@role_required('PATIENT')
def view_appointments(request):
    appointments = Appointment.objects.filter(patient=request.user).order_by('date', 'time')
    return render(request, 'accounts/patient/view_appointments.html', {
        'appointments': appointments
    })

@login_required
@role_required('PATIENT')
def cancel_appointment(request, appointment_id):
    appointment = get_object_or_404(Appointment, id=appointment_id, patient=request.user)
    
    if request.method == 'POST':
        appointment.status = 'CANCELLED'
        appointment.save()
        messages.success(request, "Rendez-vous annulé avec succès")
        return redirect('view_appointments')
    
    return render(request, 'accounts/patient/cancel_appointment.html', {
        'appointment': appointment
    })

@login_required
@role_required('PATIENT')
def reschedule_appointment(request, appointment_id):
    appointment = get_object_or_404(Appointment, id=appointment_id, patient=request.user)
    
    if request.method == 'POST':
        try:
            new_date = request.POST.get('date')
            new_time = request.POST.get('time')
            
            # Créer un nouveau rendez-vous et annuler l'ancien
            Appointment.objects.create(
                patient=request.user,
                doctor=appointment.doctor,
                date=new_date,
                time=new_time,
                reason=appointment.reason,
                custom_reason=appointment.custom_reason,
                notes=appointment.notes,
                status='SCHEDULED'
            )
            
            appointment.status = 'CANCELLED'
            appointment.save()
            
            messages.success(request, "Rendez-vous reporté avec succès!")
            return redirect('view_appointments')
            
        except Exception as e:
            messages.error(request, f"Erreur: {str(e)}")
    
    doctors = CustomUser.objects.filter(role='DOCTOR')
    return render(request, 'accounts/patient/reschedule_appointment.html', {
        'appointment': appointment,
        'doctors': doctors
    })
# views.py (ajouter ces fonctions)

@login_required
@role_required('PATIENT')
def download_medical_record(request):
    """Vue pour la page de téléchargement du dossier médical"""
    return render(request, 'accounts/patient/download_medical_record.html')

@login_required
@role_required('PATIENT')
def generate_pdf_record(request):
    """Génère un PDF du dossier médical"""
    try:
        # Récupérer les données du patient
        record = get_object_or_404(PatientRecord, patient=request.user)
        prescriptions = Prescription.objects.filter(patient=request.user)
        exams = MedicalExam.objects.filter(patient=request.user)
        
        # Créer le PDF
        buffer = BytesIO()
        p = canvas.Canvas(buffer)
        
        # En-tête
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, 800, f"Dossier médical de {request.user.get_full_name()}")
        p.setFont("Helvetica", 12)
        p.drawString(100, 780, f"Généré le {timezone.now().strftime('%d/%m/%Y')}")
        
        # Informations de base
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, 750, "Informations personnelles")
        p.setFont("Helvetica", 12)
        p.drawString(100, 730, f"Nom complet: {request.user.get_full_name()}")
        p.drawString(100, 710, f"Téléphone: {request.user.phone_number}")
        if record.primary_doctor:
            p.drawString(100, 690, f"Médecin traitant: Dr. {record.primary_doctor.get_full_name()}")
        
        # Historique médical
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, 660, "Historique médical")
        p.setFont("Helvetica", 12)
        text = p.beginText(100, 640)
        for line in record.medical_history.split('\n'):
            text.textLine(line)
        p.drawText(text)
        
        # Allergies
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, 600, "Allergies")
        p.setFont("Helvetica", 12)
        text = p.beginText(100, 580)
        for line in (record.allergies or "Aucune allergie connue").split('\n'):
            text.textLine(line)
        p.drawText(text)
        
        # Traitements actuels
        p.setFont("Helvetica-Bold", 14)
        p.drawString(100, 540, "Traitements actuels")
        p.setFont("Helvetica", 12)
        text = p.beginText(100, 520)
        for line in (record.current_medications or "Aucun traitement en cours").split('\n'):
            text.textLine(line)
        p.drawText(text)
        
        # Prescriptions
        if prescriptions.exists():
            p.setFont("Helvetica-Bold", 14)
            p.drawString(100, 480, "Prescriptions")
            y = 460
            for prescription in prescriptions:
                p.setFont("Helvetica-Bold", 12)
                p.drawString(100, y, f"{prescription.medication} - {prescription.dosage}")
                p.setFont("Helvetica", 12)
                text = p.beginText(100, y-20)
                for line in prescription.instructions.split('\n'):
                    text.textLine(line)
                p.drawText(text)
                p.drawString(100, y-40, f"Prescrit le: {prescription.created_at.strftime('%d/%m/%Y')}")
                p.drawString(100, y-60, f"Par: Dr. {prescription.doctor.get_full_name()}")
                y -= 80
                if y < 100:  # Nouvelle page si nécessaire
                    p.showPage()
                    y = 800
        
        # Examens
        if exams.exists():
            p.setFont("Helvetica-Bold", 14)
            p.drawString(100, y, "Examens médicaux")
            y -= 20
            for exam in exams:
                p.setFont("Helvetica-Bold", 12)
                p.drawString(100, y, f"{exam.get_exam_type_display()} - {exam.requested_at.strftime('%d/%m/%Y')}")
                p.setFont("Helvetica", 12)
                if exam.notes:
                    text = p.beginText(100, y-20)
                    for line in exam.notes.split('\n'):
                        text.textLine(line)
                    p.drawText(text)
                    y -= 40
                if exam.results:
                    text = p.beginText(100, y-20)
                    for line in exam.results.split('\n'):
                        text.textLine(line)
                    p.drawText(text)
                    y -= 40
                p.drawString(100, y-20, f"Demandé par: Dr. {exam.requested_by.get_full_name()}")
                y -= 40
                if y < 100:  # Nouvelle page si nécessaire
                    p.showPage()
                    y = 800
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="dossier_medical_{request.user.last_name}.pdf"'
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du PDF: {str(e)}")
        return redirect('view_medical_record')

@login_required
@role_required('PATIENT')
def generate_json_record(request):
    """Génère un fichier JSON du dossier médical"""
    try:
        # Récupérer les données du patient
        record = get_object_or_404(PatientRecord, patient=request.user)
        prescriptions = Prescription.objects.filter(patient=request.user)
        exams = MedicalExam.objects.filter(patient=request.user)
        
        # Structurer les données
        data = {
            'patient': {
                'full_name': request.user.get_full_name(),
                'phone_number': request.user.phone_number,
                'email': request.user.email,
                'primary_doctor': record.primary_doctor.get_full_name() if record.primary_doctor else None,
                'doctor_contact': record.primary_doctor.phone_number if record.primary_doctor else None
            },
            'medical_info': {
                'blood_type': record.blood_type,
                'allergies': record.allergies,
                'current_medications': record.current_medications,
                'medical_history': record.medical_history,
                'last_consultation': record.last_consultation.strftime('%Y-%m-%d') if record.last_consultation else None,
                'next_appointment': record.next_appointment.strftime('%Y-%m-%d') if record.next_appointment else None
            },
            'prescriptions': [
                {
                    'id': p.id,
                    'medication': p.medication,
                    'dosage': p.dosage,
                    'instructions': p.instructions,
                    'created_at': p.created_at.strftime('%Y-%m-%d'),
                    'doctor': p.doctor.get_full_name(),
                    'is_dispensed': p.is_dispensed,
                    'dispensed_at': p.dispensed_at.strftime('%Y-%m-%d') if p.dispensed_at else None
                } for p in prescriptions
            ],
            'exams': [
                {
                    'id': exam.id,
                    'type': exam.get_exam_type_display(),
                    'requested_at': exam.requested_at.strftime('%Y-%m-%d'),
                    'completed_at': exam.completed_at.strftime('%Y-%m-%d') if exam.completed_at else None,
                    'notes': exam.notes,
                    'results': exam.results,
                    'requested_by': exam.requested_by.get_full_name(),
                    'status': 'completed' if exam.completed_at else 'pending'
                } for exam in exams
            ],
            'generated_at': timezone.now().strftime('%Y-%m-%d %H:%M:%S'),
            'system': {
                'name': "Wergui+",
                'version': "1.0"
            }
        }
        
        response = HttpResponse(json.dumps(data, indent=2, ensure_ascii=False), content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="dossier_medical_{request.user.last_name}.json"'
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du JSON: {str(e)}")
        return redirect('view_medical_record')

@login_required
@role_required('PATIENT')
def generate_excel_record(request):
    """Génère un fichier Excel du dossier médical"""
    try:
        # Récupérer les données du patient
        record = get_object_or_404(PatientRecord, patient=request.user)
        prescriptions = Prescription.objects.filter(patient=request.user)
        exams = MedicalExam.objects.filter(patient=request.user)
        
        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        
        # Feuille d'informations
        ws_info = wb.active
        ws_info.title = "Informations"
        ws_info.append(["Dossier médical", request.user.get_full_name()])
        ws_info.append(["Date de génération", timezone.now().strftime('%d/%m/%Y %H:%M')])
        ws_info.append([])
        ws_info.append(["Informations personnelles"])
        ws_info.append(["Nom complet", request.user.get_full_name()])
        ws_info.append(["Téléphone", request.user.phone_number])
        ws_info.append(["Email", request.user.email])
        if record.primary_doctor:
            ws_info.append(["Médecin traitant", f"Dr. {record.primary_doctor.get_full_name()}"])
            ws_info.append(["Contact médecin", record.primary_doctor.phone_number])
        
        # Feuille d'informations médicales
        ws_medical = wb.create_sheet("Informations médicales")
        ws_medical.append(["Groupe sanguin", record.blood_type or "Non spécifié"])
        ws_medical.append([])
        ws_medical.append(["Allergies"])
        for line in (record.allergies or "Aucune allergie connue").split('\n'):
            ws_medical.append([line])
        ws_medical.append([])
        ws_medical.append(["Traitements actuels"])
        for line in (record.current_medications or "Aucun traitement en cours").split('\n'):
            ws_medical.append([line])
        ws_medical.append([])
        ws_medical.append(["Historique médical"])
        for line in record.medical_history.split('\n'):
            ws_medical.append([line])
        ws_medical.append([])
        ws_medical.append(["Dernière consultation", record.last_consultation.strftime('%d/%m/%Y') if record.last_consultation else "Jamais"])
        ws_medical.append(["Prochain rendez-vous", record.next_appointment.strftime('%d/%m/%Y') if record.next_appointment else "Aucun"])
        
        # Feuille des prescriptions
        if prescriptions.exists():
            ws_prescriptions = wb.create_sheet("Prescriptions")
            ws_prescriptions.append(["Date", "Médicament", "Posologie", "Instructions", "Prescrit par", "Statut"])
            for p in prescriptions:
                ws_prescriptions.append([
                    p.created_at.strftime('%d/%m/%Y'),
                    p.medication,
                    p.dosage,
                    p.instructions,
                    f"Dr. {p.doctor.get_full_name()}",
                    "Délivrée" if p.is_dispensed else "En attente"
                ])
        
        # Feuille des examens
        if exams.exists():
            ws_exams = wb.create_sheet("Examens")
            ws_exams.append(["Date demande", "Type", "Statut", "Date résultat", "Demandé par", "Notes"])
            for exam in exams:
                ws_exams.append([
                    exam.requested_at.strftime('%d/%m/%Y'),
                    exam.get_exam_type_display(),
                    "Complété" if exam.completed_at else "En attente",
                    exam.completed_at.strftime('%d/%m/%Y') if exam.completed_at else "",
                    f"Dr. {exam.requested_by.get_full_name()}",
                    exam.notes or ""
                ])
        
        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="dossier_medical_{request.user.last_name}.xlsx"'
        return response
        
    except Exception as e:
        messages.error(request, f"Erreur lors de la génération du fichier Excel: {str(e)}")
        return redirect('view_medical_record')